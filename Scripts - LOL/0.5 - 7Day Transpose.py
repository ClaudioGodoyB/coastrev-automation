import openpyxl #Variables = LOL [Property Nickname]
from datetime import datetime, timedelta
import os

# Define base paths
base_source_folder = r"/home/user/coastrev/data/extracts"
destination_file = r"C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Templates\LOL.xlsx"

# Get the date 7 days ago
seven_days_ago = (datetime.now() - timedelta(days=7)).date()

# Construct the path to the source folder from 7 days ago
source_folder = os.path.join(base_source_folder, f"Extract {seven_days_ago.strftime('%Y-%m-%d')}", "LOL")

# Check if the source folder exists
if not os.path.exists(source_folder):
    print(f"Source folder not found: {source_folder}. Skipping to next script.")
else:
    # Load destination workbook and '7Day' sheet
    dest_wb = openpyxl.load_workbook(destination_file)
    dest_sheet = dest_wb['7Day']

    # Loop through all files in the folder to find one with 'rooms_sold' in the name
    for file in os.listdir(source_folder):
        if 'rooms_sold' in file.lower() and file.endswith('.xlsx'):
            source_file = os.path.join(source_folder, file)

            # Load the source workbook
            source_wb = openpyxl.load_workbook(source_file, data_only=True)
            source_sheet = source_wb.active  # Assuming data is in the active sheet

            # Define range to copy (A2:BH432 = cols 1 to 60, rows 2 to 432)
            for i, row in enumerate(source_sheet.iter_rows(min_row=2, max_row=432, min_col=1, max_col=60), start=2):
                for j, cell in enumerate(row, start=1):
                    dest_sheet.cell(row=i, column=j, value=cell.value)

            print(f"Data copied successfully from: {file}")
            break
    else:
        print("No matching 'rooms_sold' file found in 7-day-old folder.")

    # Save changes
    dest_wb.save(destination_file)
    print("Destination file saved.")
