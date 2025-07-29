import openpyxl #Variables = CHA [Property Nickname], 'otb statistics [file keyword], line 33 for copy/paste positioning
from datetime import datetime, timedelta
import os

# Define base paths
base_source_folder = r"C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Extracts"
destination_file = r"C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Templates\CHA.xlsx"

# Get the date 7 days ago
seven_days_ago = (datetime.now() - timedelta(days=7)).date()

# Construct the path to the source folder from 7 days ago
source_folder = os.path.join(base_source_folder, f"Extract {seven_days_ago.strftime('%Y-%m-%d')}", "CHA")

# Check if the source folder exists
if not os.path.exists(source_folder):
    print(f"Source folder not found: {source_folder}. Skipping to next script.")
else:
    # Load destination workbook and '7Day' sheet
    dest_wb = openpyxl.load_workbook(destination_file)
    dest_sheet = dest_wb['7Day']

    # Loop through all files in the folder to find one with 'otb statistics' in the name
    for file in os.listdir(source_folder):
        if 'otb statistics' in file.lower() and file.endswith('.xlsx'):
            source_file = os.path.join(source_folder, file)

            # Load the source workbook
            source_wb = openpyxl.load_workbook(source_file, data_only=True)
            source_sheet = source_wb.active  # Assuming data is in the active sheet

            # Define range to copy (A2:BH432 = cols 1 to 60, rows 2 to 432)
            for i, row in enumerate(source_sheet.iter_rows(min_row=1, max_row=432, min_col=1, max_col=60), start=1):
                for j, cell in enumerate(row, start=1):
                    dest_sheet.cell(row=i, column=j, value=cell.value)

            print(f"Data copied successfully from: {file}")
            break
    else:
        print("No matching 'otb statistics' file found in 7-day-old folder.")

    # Save changes
    dest_wb.save(destination_file)
    print("Destination file saved.")
