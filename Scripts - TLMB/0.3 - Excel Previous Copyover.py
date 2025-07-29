import openpyxl # Variable items = TLMB (Property nickname), line 13 range (columns)

try:
    # Load the Excel file
    file_path = r"C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Templates\TLMB.xlsx"
    wb = openpyxl.load_workbook(file_path)

    # Get the 'Today' and 'Yesterday' sheets
    current_sheet = wb["Today"]
    prev_sheet = wb["Yesterday"]

    # Copy columns A:Z from 'Today' to 'Yesterday'
    for col in range(1, 60):  # Columns A:BH represented as 1 to 60
        for row in range(1, current_sheet.max_row + 1):
            cell_value = current_sheet.cell(row=row, column=col).value
            prev_sheet.cell(row=row, column=col).value = cell_value

    # Save the changes and close the workbook
    wb.save(file_path)
    wb.close()

    # Print success message
    print("Data Copied Successfully")

except Exception as e:
    # Print failure message
    print(f"Failed to copy data: {e}")
