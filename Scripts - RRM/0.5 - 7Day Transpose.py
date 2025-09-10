import openpyxl
from datetime import datetime, timedelta
import os

# Define base paths
base_source_folder = r"/home/user/coastrev/data/extracts"
destination_file = r"C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Templates\RRM.xlsx"

# Calculate the date for 7 days ago
seven_days_ago = datetime.now().date() - timedelta(days=7)

# Construct the folder path for 7 days ago's extract, including the 'RRM' subfolder
source_folder = os.path.join(base_source_folder, f"Extract {seven_days_ago.strftime('%Y-%m-%d')}", "RRM")

# Check if the source folder exists
if not os.path.exists(source_folder):
    print(f"The source folder '{source_folder}' does not exist. No data was copied.")
else:
    # Load destination workbook
    dest_wb = openpyxl.load_workbook(destination_file)
    dest_sheet = dest_wb['7Day']  # Set the target tab to '7Day'

    source_files_found = False  # Track if any files were processed

    # Loop through all files in the 'RRM' subfolder
    for file in os.listdir(source_folder):
        if file.endswith('.xlsx') and "occupancy" in file.lower() and "occupancy-by-room-type" not in file.lower():
            source_files_found = True
            source_file = os.path.join(source_folder, file)

            try:
                # Load source workbook and data
                source_wb = openpyxl.load_workbook(source_file, data_only=True)
                source_sheet = source_wb.active  # Assuming data is in the active sheet

                # Copy the data from source (range B13:AI438) to the destination, starting at cell D11 in '7Day' tab
                data_range = source_sheet['B13:AI438']
                start_row = 11  # Starting row for destination
                start_col = 4   # Starting column D in destination

                for i, row in enumerate(data_range, start=0):
                    for j, cell in enumerate(row, start=0):
                        dest_sheet.cell(row=start_row + i, column=start_col + j, value=cell.value)

                print(f"Data copied successfully from {file}.")
            except Exception as e:
                print(f"An error occurred while processing the file {file}: {e}")

    # Save the updated destination file if any files were processed
    if source_files_found:
        dest_wb.save(destination_file)
        print("All data copied successfully.")
    else:
        print(f"No valid source files found in '{source_folder}'.")
