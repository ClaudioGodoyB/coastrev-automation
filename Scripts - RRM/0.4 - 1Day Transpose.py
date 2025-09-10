import openpyxl
from datetime import datetime, timedelta
import os

# Define base paths
base_source_folder = r"/home/user/coastrev/data/extracts"
destination_file = r"C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Templates\RRM.xlsx"

# Get today's date
today = datetime.now().date()

# Construct the folder path for today's extract, including the 'RRM' subfolder
source_folder = os.path.join(base_source_folder, f"Extract {today.strftime('%Y-%m-%d')}", "RRM")

# Load destination workbook
dest_wb = openpyxl.load_workbook(destination_file)
dest_sheet = dest_wb['Today']

# Loop through all files in the 'RRM' subfolder
for file in os.listdir(source_folder):
    if file.endswith('.xlsx') and "occupancy" in file.lower() and "occupancy-by-room-type" not in file.lower():
        source_file = os.path.join(source_folder, file)

        # Load source workbook and data
        source_wb = openpyxl.load_workbook(source_file, data_only=True)
        source_sheet = source_wb.active  # Assuming data is in the active sheet

        # Copy the data from source (range B13:AI438) to the destination, starting at cell D11 in 'Today' tab
        data_range = source_sheet['B13:AI438']
        start_row = 11  # Starting row for destination
        start_col = 4   # Starting column D in destination

        for i, row in enumerate(data_range, start=0):
            for j, cell in enumerate(row, start=0):
                dest_sheet.cell(row=start_row + i, column=start_col + j, value=cell.value)

        print(f"Data copied successfully from {file}.")

# Save the updated destination file after all files are processed
dest_wb.save(destination_file)
print("All data copied successfully.")
