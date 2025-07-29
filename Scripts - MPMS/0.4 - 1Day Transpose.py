import os
from openpyxl import load_workbook
from datetime import datetime

try:
    # Determine today's date
    today_str = datetime.now().strftime("%Y-%m-%d")
    extract_folder = rf"C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Extracts\Extract {today_str}\MPMS"

    # Find the source file with "booking" in the name (case-insensitive)
    source_file = None
    for file in os.listdir(extract_folder):
        if "booking" in file.lower() and file.lower().endswith(".xlsx"):
            source_file = os.path.join(extract_folder, file)
            break

    if not source_file:
        print("❌ No file containing 'booking' was found in today's extract folder.")
    else:
        # Load source workbook and read A:Z
        src_wb = load_workbook(source_file)
        src_ws = src_wb.active  # Assuming data is in the first sheet

        # Load destination workbook
        dest_path = r"C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Templates\MPMS.xlsx"
        dest_wb = load_workbook(dest_path)
        dest_ws = dest_wb["Today"]

        # Clear existing B:AA data before pasting new data
        for row in range(1, dest_ws.max_row + 1):
            for col in range(2, 28):  # Columns B (2) to AA (27)
                dest_ws.cell(row=row, column=col).value = None

        # Copy A:Z from source and paste into B:AA in destination
        for row_idx, row in enumerate(src_ws.iter_rows(min_col=1, max_col=26), start=1):
            for col_idx, cell in enumerate(row, start=2):  # Start at column 2 (B)
                dest_ws.cell(row=row_idx, column=col_idx, value=cell.value)

        dest_wb.save(dest_path)
        print(f"✅ Data from '{os.path.basename(source_file)}' successfully copied to 'Today' tab in MPMS.xlsx")

except Exception as e:
    print(f"❌ Error during execution: {e}")
