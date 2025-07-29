import os
from openpyxl import load_workbook
from datetime import datetime, timedelta

try:
    # Date from 7 days ago
    target_date = datetime.now() - timedelta(days=7)
    target_str = target_date.strftime("%Y-%m-%d")
    extract_folder = rf"C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Extracts\Extract {target_str}\MPMS"

    # If the folder doesn't exist, skip
    if not os.path.exists(extract_folder):
        print(f"⚠️ Extract folder from 7 days ago does not exist: {extract_folder}")
    else:
        # Look for file with 'booking' in name
        source_file = None
        for file in os.listdir(extract_folder):
            if "booking" in file.lower() and file.lower().endswith(".xlsx"):
                source_file = os.path.join(extract_folder, file)
                break

        # If no file found, skip
        if not source_file:
            print("⚠️ No 'booking' file found in the extract folder from 7 days ago.")
        else:
            # Load source workbook and read A:Z
            src_wb = load_workbook(source_file)
            src_ws = src_wb.active  # Use the first sheet

            # Load destination workbook
            dest_path = r"C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Templates\MPMS.xlsx"
            dest_wb = load_workbook(dest_path)
            dest_ws = dest_wb["7Day"]

            # Clear columns B:AA
            for row in range(1, dest_ws.max_row + 1):
                for col in range(2, 28):  # B to AA
                    dest_ws.cell(row=row, column=col).value = None

            # Copy from A:Z → B:AA
            for row_idx, row in enumerate(src_ws.iter_rows(min_col=1, max_col=26), start=1):
                for col_idx, cell in enumerate(row, start=2):
                    dest_ws.cell(row=row_idx, column=col_idx, value=cell.value)

            dest_wb.save(dest_path)
            print(f"✅ Data from '{os.path.basename(source_file)}' copied to '7Day' tab in MPMS.xlsx")

except Exception as e:
    print(f"❌ Error occurred but continuing gracefully: {e}")
