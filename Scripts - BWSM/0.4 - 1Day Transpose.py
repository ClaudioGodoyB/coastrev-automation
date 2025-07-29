import os
from datetime import datetime
from openpyxl import load_workbook

def transpose_data_to_template():
    # Build today's date string for path
    today_str = datetime.today().strftime("%Y-%m-%d")
    today_display = datetime.today().strftime("%d_%b_%Y")  # e.g. 21_Apr_2025

    # File paths
    source_file = fr'C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Extracts\Extract {today_str}\BWSM\05232_metrics_{today_display}_*.xlsx'
    destination_file = r'C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Templates\BWSM.xlsx'

    # Resolve the actual source filename if wildcards exist
    folder = os.path.dirname(source_file)
    prefix = f"05232_metrics_{today_display}"
    source_actual = None
    for fname in os.listdir(folder):
        if fname.startswith(prefix) and fname.endswith(".xlsx"):
            source_actual = os.path.join(folder, fname)
            break

    if not source_actual or not os.path.exists(source_actual):
        print("Source file not found.")
        return

    if not os.path.exists(destination_file):
        print("Destination file not found.")
        return

    # Load workbooks
    src_wb = load_workbook(source_actual)
    src_ws = src_wb.active  # Assume first sheet

    dest_wb = load_workbook(destination_file)
    dest_ws = dest_wb['Today']

    # Build mapping from date in destination (column A) to its row number
    dest_date_row_map = {}
    for row in range(2, dest_ws.max_row + 1):
        cell = dest_ws[f'A{row}'].value
        if isinstance(cell, datetime):
            dest_date_row_map[cell.date()] = row
        elif cell:
            try:
                dest_date_row_map[datetime.strptime(str(cell), "%Y-%m-%d").date()] = row
            except:
                pass

    # Loop through source rows and paste values based on date in column C
    for row in range(2, src_ws.max_row + 1):
        date_cell = src_ws[f'C{row}'].value
        if isinstance(date_cell, datetime):
            src_date = date_cell.date()
        elif date_cell:
            try:
                src_date = datetime.strptime(str(date_cell), "%Y-%m-%d").date()
            except:
                continue
        else:
            continue

        if src_date in dest_date_row_map:
            dest_row = dest_date_row_map[src_date]

            # Copy A:AZ from source to B:BA in destination
            for col_offset in range(52):  # 52 columns = A:AZ
                src_val = src_ws.cell(row=row, column=1 + col_offset).value  # A=1
                dest_ws.cell(row=dest_row, column=2 + col_offset).value = src_val  # B=2

    # Save destination workbook
    dest_wb.save(destination_file)
    print("Data transfer completed successfully.")

if __name__ == '__main__':
    transpose_data_to_template()
