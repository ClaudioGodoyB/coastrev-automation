import os #Variable items = BWSM (Property nickname), 27231 (Expedia Property ID)
import win32com.client as win32

try:
    from openpyxl import load_workbook
    from datetime import datetime

    # Get today's date in the required format
    today_str = datetime.today().strftime('%Y-%m-%d')

    # Paths
    source_folder = fr"C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Extracts\Extract {today_str}"
    destination_path = r"C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Templates\BWSM.xlsx"

    # Find the source file with the [Expedia Property ID #] in the title
    source_file = None
    for file_name in os.listdir(source_folder):
        if '27231' in file_name.lower():
            source_file = os.path.join(source_folder, file_name)
            break

    if not source_file:
        raise FileNotFoundError(f"No file with 'expedia' in the title found in the source folder: {source_folder}")

    # Open the source file using win32com.client
    excel = win32.Dispatch('Excel.Application')
    workbook_source = excel.Workbooks.Open(source_file)
    sheet_source = workbook_source.ActiveSheet

    # Copy the value from cell A2 in the source file
    cell_a2_value = sheet_source.Cells(2, 1).Value

    # Read the data from rows 12 through 39
    data = []
    for row in range(12, 40):  # Excel rows are 1-indexed
        row_data = []
        for col in range(1, sheet_source.UsedRange.Columns.Count + 1):  # Copy all columns in the specified rows
            row_data.append(sheet_source.Cells(row, col).Value)
        data.append(row_data)

    # Close the source file
    workbook_source.Close(SaveChanges=False)

    # Load the destination workbook and select the 'Rates' sheet
    wb_dest = load_workbook(destination_path)
    sheet_dest = wb_dest['Rates']

    # Paste the copied data into the destination file starting from row 12
    for i, row_data in enumerate(data, start=12):
        for j, value in enumerate(row_data, start=1):
            sheet_dest.cell(row=i, column=j).value = value

    # Paste the value from cell A2 into cell A2 of the 'Rates' sheet
    sheet_dest.cell(row=2, column=1).value = cell_a2_value

    # Save the destination file
    wb_dest.save(destination_path)
    wb_dest.close()

    print("Data copied successfully.")

except Exception as e:
    print(f"Error occurred: {e}")
