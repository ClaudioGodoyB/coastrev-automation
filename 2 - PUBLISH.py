from openpyxl import load_workbook
import win32com.client as win32
from os.path import join, exists
from datetime import datetime

# Function to send email using Outlook with attachment and custom sender
def send_email_with_attachment(to_email, subject, body, attachment_path, from_email):
    outlook = win32.Dispatch('Outlook.Application')
    mail = outlook.CreateItem(0)

    # Set sender account
    accounts = outlook.Session.Accounts
    for account in accounts:
        if account.SmtpAddress.lower() == from_email.lower():
            mail._oleobj_.Invoke(*(64209, 0, 8, 0, account))
            break
    else:
        print(f"Sender email account '{from_email}' not found in Outlook. Email not sent.")
        return

    mail.To = to_email
    mail.Subject = subject
    mail.HTMLBody = body  # Use HTMLBody for HTML content

    # Attach the file if it exists
    if exists(attachment_path):
        mail.Attachments.Add(attachment_path)
        mail.Send()
    else:
        print(f"Attachment file not found at: {attachment_path}. Email not sent to {to_email}.")

# Load the Excel workbook
file_path = r'C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Templates\!Daily Detail Roster & Distro.xlsx'
wb = load_workbook(file_path, data_only=True)
sheet = wb.active

# Get headers and data from the Excel table
headers = [cell.value for cell in sheet[1]]
data = [[cell.value for cell in row] for row in sheet.iter_rows(min_row=2)]

# Path for HTML files
current_date = datetime.now().strftime("%Y-%m-%d")
html_base_path = fr'C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Details\Daily Detail {current_date}\Misc'

# Iterate through each row and send email with HTML content and attachment
for row in data:
    sent_today = row[headers.index('SENT TODAY')] if 'SENT TODAY' in headers else None
    if not sent_today or str(sent_today).strip().upper() != 'NO':
        continue  # Skip rows not marked as 'NO'

    inn_code = row[headers.index('INN CODE')] if 'INN CODE' in headers else None
    subject_cell = row[headers.index('SUBJECT')] if 'SUBJECT' in headers else None
    property_name = row[headers.index('PROPERTY NAME')] if 'PROPERTY NAME' in headers else None
    folder = row[headers.index('FOLDER')] if 'FOLDER' in headers else None
    attachment_file = row[headers.index('FILE')] if 'FILE' in headers else None
    email_list = row[headers.index('EMAIL')] if 'EMAIL' in headers else None
    send_email = row[headers.index('SEND EMAIL')] if 'SEND EMAIL' in headers else None
    pickup_html = row[headers.index('PICKUP HTML')] if 'PICKUP HTML' in headers else None

    # Ensure required fields are not empty
    if inn_code and email_list and subject_cell and property_name and folder and attachment_file and send_email and pickup_html:
        if isinstance(subject_cell, str) and subject_cell.startswith('='):
            formula_result = sheet.cell(row=data.index(row) + 2, column=headers.index('SUBJECT') + 1).value
            subject = formula_result if formula_result is not None else subject_cell
        else:
            subject = subject_cell

        base_path = r"/home/user/coastrev/data/daily_details"
        attachment_path = join(base_path, folder, f"{attachment_file}.xlsx")

        html_file_path = join(html_base_path, f"{pickup_html}.html")
        if exists(html_file_path):
            with open(html_file_path, 'r', encoding='utf-8') as html_file:
                html_content = html_file.read()
            body_message = f"<p>Attached is the daily pickup report for {property_name}.</p>{html_content}"
        else:
            print(f"HTML file not found at: {html_file_path}. Email not sent for INN CODE: {inn_code}.")
            continue

        if exists(attachment_path):
            for email in email_list.split(','):
                send_email_with_attachment(email.strip(), subject, body_message, attachment_path, send_email.strip())
        else:
            print(f"Attachment file not found at: {attachment_path}. No email sent for INN CODE: {inn_code}.")
