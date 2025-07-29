import os  # Variables = BWOF [Property Nickname], 10437 [Property Brand ID]
from datetime import datetime, timedelta
from openpyxl import load_workbook

def transpose_7day_data_to_template():
    try:
        # Get date 7 days ago
        target_date = datetime.today() - timedelta(days=7)
        folder_date_str = target_date.strftime("%Y-%m-%d")
        file_date_str = f"{target_date.day}_{target_date.strftime('%b_%Y')}"  # e.g. 6_May_2025

        # File paths
        source_dir = fr'C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Extracts\Extract {folder_date_str}\BWOF'
        destination_file = r'C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Templates\BWOF.xlsx'

        # Check if source folder exists
        if not os.path.exists(source_dir):
            print(f"Source folder not found: {source_dir}. Skipping.")
            return

        # Find the source file in that folder
        source_file = None
        for fname in os.listdir(source_dir):
            if fname.startswith(f'10437_metrics_{file_date_str}') and fname.endswith('.xlsx'):
                source_file = os.path.join(source_dir, fname)
                break

        if not source_file or not os.path.exists(source_file):
            print("Source file not found from 7 days ago. Skipping.")
            return

        if not os.path.exists(destination_file):
            print("Destination file not found. Skipping.")
            return

        # Load workbooks
        src_wb = load_workbook(source_file)
        src_ws = src_wb.active

        dest_wb = load_workbook(destination_file)
        dest_ws = dest_wb['7Day']

        # Build mapping from date in destination column A → row number
        dest_date_row_map = {}
        for row in range(2, dest_ws.max_row + 1):
            cell = dest_ws[f'A{row}'].value
            if isinstance(cell, datetime):
                dest_date_row_map[cell.date()] = row
            elif cell:
                try:
                    dest_date_row_map[datetime.strptime(str(cell), "%Y-%m-%d").date()] = row
                except:
                    pass

        # Loop through source and match against dates in column C
        for row in range(2, src_ws.max_row + 1):
            date_cell = src_ws[f'C{row}'].value
            if isinstance(date_cell, datetime):
                src_date = date_cell.date()
            elif date_cell:
                try:
                    src_date = datetime.strptime(str(date_cell), "%Y-%m-%d").date()
                except:
                    continue
            else:
                continue

            if src_date in dest_date_row_map:
                dest_row = dest_date_row_map[src_date]
                for col_offset in range(52):  # A:AZ → 52 columns
                    src_val = src_ws.cell(row=row, column=1 + col_offset).value
                    dest_ws.cell(row=dest_row, column=2 + col_offset).value = src_val  # paste into B:BA

        dest_wb.save(destination_file)
        print("7-day data successfully copied to '7Day' tab.")

    except Exception as e:
        print(f"An error occurred, but script will continue. Error: {e}")

if __name__ == '__main__':
    transpose_7day_data_to_template()
