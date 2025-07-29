import openpyxl  # Variable items = MSI (Property nickname)
from datetime import datetime, timedelta
import os

# Define base paths
base_source_folder = r"C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Extracts"
destination_file = r"C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Templates\MSI.xlsx"

# Get today's date
today = datetime.now().date()

# Calculate the date 7 days ago for the folder and 8 days ago for the file name
seven_days_ago = today - timedelta(days=7)
eight_days_ago = today - timedelta(days=8)

# Construct the folder path for the 7 days ago extract, including the 'MSI' subfolder
source_folder = os.path.join(base_source_folder, f"Extract {seven_days_ago.strftime('%Y-%m-%d')}", "MSI")

# Initialize source file variable
source_file = None

# Find the source file in the folder with the date 8 days ago
if os.path.exists(source_folder):
    for file in os.listdir(source_folder):
        if file.endswith('.xlsx') and eight_days_ago.strftime('%d-%b-%Y') in file:
            source_file = os.path.join(source_folder, file)
            break

if not source_file:
    print(f"No source file found for {eight_days_ago.strftime('%d-%b-%Y')} in the {seven_days_ago.strftime('%Y-%m-%d')} 'MSI' subfolder.")
else:
    try:
        # Extract date from filename (this assumes the date in the filename is still relevant)
        filename = source_file.split('\\')[-1]
        date_str = filename.split('_')[2]  # e.g., "11-Aug-2024"
        date_to_find = datetime.strptime(date_str, '%d-%b-%Y').date()  # Convert to date object

        # Load source workbook and data
        source_wb = openpyxl.load_workbook(source_file, data_only=True)
        source_sheet = source_wb.active  # Assuming data is in the active sheet

        # Load destination workbook with data_only to get the calculated values
        dest_wb = openpyxl.load_workbook(destination_file, data_only=True)
        dest_sheet = dest_wb['7Day']  # Ensure we're on the '7Day' tab

        # Locate the row with the corresponding date in the destination sheet
        date_column = 3  # Column C is the 3rd column
        start_row = None

        for row in dest_sheet.iter_rows(min_row=1, max_row=dest_sheet.max_row, min_col=date_column, max_col=date_column):
            cell = row[0]
            if isinstance(cell.value, datetime) and cell.value.date() == date_to_find:
                start_row = cell.row
                break

        if start_row is None:
            print("Date not found in destination sheet.")
        else:
            # Load destination workbook again, but with formulas visible, to paste data correctly
            dest_wb = openpyxl.load_workbook(destination_file)
            dest_sheet = dest_wb['7Day']  # Ensure we're still on the '7Day' tab

            # Define the range of data to copy from source
            data_range = source_sheet['A11:X375']

            # Paste the data into the destination sheet starting at the adjacent column to the found date
            for i, row in enumerate(data_range, start=0):
                for j, cell in enumerate(row, start=0):
                    dest_sheet.cell(row=start_row + i, column=j + 4, value=cell.value)

            # Save the updated destination file
            dest_wb.save(destination_file)
            print("Data copied successfully.")
    except Exception as e:
        print(f"An error occurred while processing the source file: {e}")
