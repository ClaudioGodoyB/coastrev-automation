import openpyxl #Variable items = MSI (Property nickname)
from datetime import datetime, timedelta
import os

# Check if today is Monday; if not, exit the script
if datetime.now().weekday() != 0:
    print("The script only runs on Mondays.")
    exit()

# Define base paths
base_source_folder = r"/home/user/coastrev/data/extracts"
destination_file = r"C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Templates\MSI.xlsx"

# Get today's date (adjusted to refer to yesterday as originally scripted)
today = datetime.now().date() - timedelta(days=2)

# Construct the folder path for today's extract, including the 'MSI' subfolder
source_folder = os.path.join(base_source_folder, f"Extract {today.strftime('%Y-%m-%d')}", "MSI")

# Calculate the date one day prior for the file name
previous_day = today - timedelta(days=1)

# Find the source file in the 'MSI' subfolder
for file in os.listdir(source_folder):
    if file.endswith('.xlsx') and previous_day.strftime('%d-%b-%Y') in file:
        source_file = os.path.join(source_folder, file)
        break
else:
    print(f"No source file found for {previous_day.strftime('%d-%b-%Y')} in today's 'MSI' subfolder.")
    exit()

# Extract date from filename (this assumes the date in the filename is still relevant)
filename = source_file.split('\\')[-1]
date_str = filename.split('_')[2]  # "18-Aug-2024"
date_to_find = datetime.strptime(date_str, '%d-%b-%Y').date()  # Convert to date object

# Load source workbook and data
source_wb = openpyxl.load_workbook(source_file, data_only=True)
source_sheet = source_wb.active  # Assuming data is in the active sheet

# Load destination workbook with data_only to get the calculated values
dest_wb = openpyxl.load_workbook(destination_file, data_only=True)
dest_sheet = dest_wb['Today']

# Locate the row with the corresponding date in the destination sheet
date_column = 3  # Column C is the 3rd column
start_row = None

for row in dest_sheet.iter_rows(min_row=1, max_row=dest_sheet.max_row, min_col=date_column, max_col=date_column):
    cell = row[0]
    if isinstance(cell.value, datetime) and cell.value.date() == date_to_find:
        start_row = cell.row
        break

if start_row is None:
    print("Date not found in destination sheet.")
else:
    # Load destination workbook again, but with formulas visible, to paste data correctly
    dest_wb = openpyxl.load_workbook(destination_file)
    dest_sheet = dest_wb['Today']

    # Define the range of data to copy from source
    data_range = source_sheet['A11:X11']

    # Paste the data into the destination sheet starting at the adjacent column to the found date
    for i, row in enumerate(data_range, start=0):
        for j, cell in enumerate(row, start=0):
            dest_sheet.cell(row=start_row + i, column=j + 4, value=cell.value)

    # Save the updated destination file
    dest_wb.save(destination_file)
    print("Data copied successfully.")
