import openpyxl # HBSB [Property nickname],'Rooms Sold' [Extract file name snippet]
from datetime import datetime
import os

# Define base paths
base_source_folder = r"/home/user/coastrev/data/extracts"
destination_file = r"C:\Users\johnj\Desktop\CoastRev\Reporting\Daily Templates\HBSB.xlsx"

# Get today's date
today = datetime.now().date()

# Construct today's extract path
source_folder = os.path.join(base_source_folder, f"Extract {today.strftime('%Y-%m-%d')}", "HBSB")

# Load destination workbook and 'Today' sheet
dest_wb = openpyxl.load_workbook(destination_file)
dest_sheet = dest_wb['Today']

# Loop through all files in the folder to find one with 'Rooms_Sold' in the name
for file in os.listdir(source_folder):
    if 'rooms_sold' in file.lower() and '_1.xlsx' in file.lower() and file.endswith('.xlsx'):
        source_file = os.path.join(source_folder, file)

        # Load the source workbook
        source_wb = openpyxl.load_workbook(source_file, data_only=True)
        source_sheet = source_wb.active  # Assuming data is in the active sheet

        # Define range to copy (A2:BH432 = cols 1 to 60, rows 2 to 432)
        for i, row in enumerate(source_sheet.iter_rows(min_row=2, max_row=432, min_col=1, max_col=60), start=2):
            for j, cell in enumerate(row, start=1):
                dest_sheet.cell(row=i, column=j, value=cell.value)

        print(f"Data copied successfully from: {file}")
        break
else:
    print("No matching 'Rooms_Sold' file found in today's folder.")

# Save changes
dest_wb.save(destination_file)
print("Destination file saved.")
